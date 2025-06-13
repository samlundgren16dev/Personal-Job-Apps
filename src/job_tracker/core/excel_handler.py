"""
Excel handler module for job application tracker.

This module provides functionality to read, write, and manage job application data
stored in Excel format with caching and backup capabilities.
"""

import os
import json
from pathlib import Path
from datetime import datetime

import openpyxl

from .constants import EXCEL_FILE, HEADERS, GREEN, YELLOW, RED, CYAN, RESET

CONFIG_DIR = Path(__file__).parent / "config"
CONFIG_FILE = CONFIG_DIR / "user_config.json"

# Add workbook cache
_workbook_cache = None
_cache_dirty = False


def find_file(filename, search_path=None):
    """Find a file in the filesystem starting from search_path."""
    print(f"{CYAN}Beginning File Search for {filename}{RESET}")
    if search_path is None:
        search_path = Path.home()
    for root, dirs, files in os.walk(search_path):
        if filename in files:
            found_file = os.path.join(root, filename)
            print(f"{GREEN}Found file: {found_file}{RESET}")
            return found_file
    print(f"{RED}File not found during disk search.{RESET}")
    return None


def save_config(excel_path):
    """Save Excel file path to configuration file."""
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(CONFIG_FILE, "w", encoding='utf-8') as f:
        json.dump({"excel_path": excel_path}, f)
    print(f"{GREEN}Successfully saved the excel file path to config file{RESET}")


def load_config():
    """Load Excel file path from configuration file."""
    if CONFIG_FILE.exists():
        print(f"{YELLOW}Attempting to load from config...{RESET}")
        try:
            with open(CONFIG_FILE, "r", encoding='utf-8') as f:
                content = f.read().strip()
                if not content:
                    print(f"{RED}Config file is empty.{RESET}")
                    return None
                data = json.loads(content)
                return data.get("excel_path")
        except (json.JSONDecodeError, OSError):
            print(f"{RED}Failed to load from config (invalid or unreadable).{RESET}")
            return None
    print(f"{RED}Config file does not exist or can't be accessed.{RESET}")
    return None


def get_excel_path():
    """Get Excel file path from config or find it on disk."""
    # Try loading from config first
    cached_path = load_config()
    if cached_path and os.path.exists(cached_path):
        print(f"{GREEN}Loaded file from config: {cached_path}{RESET}")
        return cached_path

    # Try to find the file on disk
    found_path = find_file(EXCEL_FILE)
    if found_path:
        print(f"{GREEN}Loaded file from disk search: {found_path}{RESET}")
        save_config(found_path)
        return found_path

    # Fallback
    default_path = os.path.join(os.getcwd(), EXCEL_FILE)
    print(f"{RED}Excel file not found. Creating new one at: {default_path}{RESET}")
    save_config(default_path)
    return default_path


EXCEL_PATH = get_excel_path()


def _get_workbook():
    """Get cached workbook or load from file."""
    global _workbook_cache
    if _workbook_cache is None:
        try:
            _workbook_cache = openpyxl.load_workbook(EXCEL_PATH)
        except FileNotFoundError:
            init_excel()
            _workbook_cache = openpyxl.load_workbook(EXCEL_PATH)
    return _workbook_cache


def _save_workbook():
    """Save workbook only if changes were made."""
    global _cache_dirty
    if _cache_dirty and _workbook_cache:
        _workbook_cache.save(EXCEL_PATH)
        _cache_dirty = False


def flush_changes():
    """Force save any pending changes."""
    _save_workbook()


def init_excel():
    """Initialize Excel file with headers if it doesn't exist."""
    global _workbook_cache, _cache_dirty
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)


def save_to_excel(row_data):
    """Save a single row of data to Excel."""
    global _cache_dirty
    wb = _get_workbook()
    ws = wb.active
    ws.append(row_data)
    _cache_dirty = True
    _save_workbook()  # Still save immediately for data safety


def delete_from_excel(values):
    """Delete a row from Excel based on matching values."""
    global _cache_dirty
    wb = _get_workbook()
    ws = wb.active

    found = False
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def normalize(val):
            return str(val).strip().replace('\n', '').replace('\r', '') if val is not None else ""

        compare_values = [normalize(val) for val in values]
        row_values = [normalize(cell.value) for cell in row[:len(values)]]  # Only compare up to values length

        # Exact match required - all values must match
        if len(row_values) == len(compare_values) and row_values == compare_values:
            ws.delete_rows(idx)
            found = True
            _cache_dirty = True
            print("Successfully Deleted Row")
            break

    _save_workbook()
    return found


def get_all_applications():
    """Get all job applications from Excel."""
    wb = _get_workbook()
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


def update_excel_row(old_values, new_values):
    """Update an Excel row by replacing old values with new values."""
    global _cache_dirty
    wb = _get_workbook()
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value if cell.value is not None else "" for cell in row]
        if row_values == [v if v is not None else "" for v in old_values]:
            for idx, val in enumerate(new_values):
                row[idx].value = val
            _cache_dirty = True
            break
    _save_workbook()


def batch_save_to_excel(rows_data):
    """Save multiple rows at once."""
    global _cache_dirty
    wb = _get_workbook()
    ws = wb.active
    for row_data in rows_data:
        ws.append(row_data)
    _cache_dirty = True
    _save_workbook()


def clear_cache():
    """Clear the workbook cache (useful for testing or memory cleanup)."""
    global _workbook_cache, _cache_dirty
    _workbook_cache = None
    _cache_dirty = False


def export_to_csv(output_path):
    """Export job applications to CSV file."""
    import csv
    try:
        applications = get_all_applications()
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(HEADERS)  # Write headers
            writer.writerows(applications)
        print(f"{GREEN}Successfully exported to {output_path}{RESET}")
        return True
    except Exception as e:
        print(f"{RED}Export failed: {e}{RESET}")
        return False


def import_from_csv(input_path):
    """Import job applications from CSV file."""
    import csv
    try:
        with open(input_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip header row

            rows_to_import = []
            for row in reader:
                # Pad row if it's shorter than expected
                while len(row) < len(HEADERS):
                    row.append("")
                # Trim if longer
                rows_to_import.append(row[:len(HEADERS)])

            batch_save_to_excel(rows_to_import)
            imported_count = len(rows_to_import)
            print(f"{GREEN}Successfully imported {imported_count} jobs from {input_path}{RESET}")
            return imported_count
    except Exception as e:
        print(f"{RED}Import failed: {e}{RESET}")
        return 0


def backup_data():
    """Create a backup of the current data (overwrites existing backup)."""
    import shutil

    try:
        backup_filename = "job_applications_backup.xlsx"
        backup_path = Path(EXCEL_PATH).parent / "backups" / backup_filename

        # Create backups directory if it doesn't exist
        backup_path.parent.mkdir(exist_ok=True)

        # Force save current changes
        flush_changes()

        # Copy file (will overwrite existing backup)
        shutil.copy2(EXCEL_PATH, backup_path)
        print(f"{GREEN}Backup created/updated: {backup_path}{RESET}")
        return str(backup_path)
    except Exception as e:
        print(f"{RED}Backup failed: {e}{RESET}")
        return None


def get_statistics():
    """Get basic statistics about job applications."""
    try:
        applications = get_all_applications()
        if not applications:
            return {"total": 0, "companies": [], "locations": []}

        companies = {}
        locations = {}

        # Count companies and locations
        for app in applications:
            if len(app) >= 3:  # Ensure we have company data
                company = str(app[2]).strip() if app[2] else "Unknown"
                location = str(app[4]).strip() if len(app) > 4 and app[4] else "Unknown"

                companies[company] = companies.get(company, 0) + 1
                locations[location] = locations.get(location, 0) + 1

        top_companies = sorted(companies.items(), key=lambda x: x[1], reverse=True)[:10]
        top_locations = sorted(locations.items(), key=lambda x: x[1], reverse=True)[:10]

        return {
            "total": len(applications),
            "companies": top_companies,
            "locations": top_locations,
            "recent_count": len([app for app in applications
                               if len(app) > 0 and str(app[0]) >= datetime.now().strftime("%Y-%m-%d")])
        }
    except Exception as e:
        print(f"{RED}Statistics calculation failed: {e}{RESET}")
        return {"total": 0, "companies": [], "locations": []}
