import openpyxl
import os
import json
from pathlib import Path
from constants import EXCEL_FILE, HEADERS, GREEN, YELLOW, RED, CYAN, RESET

CONFIG_DIR = Path(__file__).parent / "config"
CONFIG_FILE = CONFIG_DIR / "user_config.json"

def find_file(filename, search_path=None):
    print(f"{CYAN}Beginning File Search for {filename}{RESET}")
    if search_path is None:
        search_path = Path.home()
    for root, dirs, files in os.walk(search_path):
        if filename in files:
            found_file = os.path.join(root, filename)
            print(f"{GREEN}Found file: {found_file}{RESET}")
            return found_file
    print(f"{RED}File not found during disk search.{RESET}")
    return None

def save_config(excel_path):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump({"excel_path": excel_path}, f)
    print(f"{GREEN}Successfully saved the excel file path to config file{RESET}")

def load_config():
    if CONFIG_FILE.exists():
        print(f"{YELLOW}Attempting to load from config...{RESET}")
        try:
            with open(CONFIG_FILE, "r") as f:
                content = f.read().strip()
                if not content:
                    print(f"{RED}Config file is empty.{RESET}")
                    return None
                data = json.loads(content)
                return data.get("excel_path")
        except (json.JSONDecodeError, OSError):
            print(f"{RED}Failed to load from config (invalid or unreadable).{RESET}")
            return None
    print(f"{RED}Config file does not exist or can't be accessed.{RESET}")
    return None

def get_excel_path():
    # Try loading from config first
    cached_path = load_config()
    if cached_path and os.path.exists(cached_path):
        print(f"{GREEN}Loaded file from config: {cached_path}{RESET}")
        return cached_path

    # Try to find the file on disk
    found_path = find_file(EXCEL_FILE)
    if found_path:
        print(f"{GREEN}Loaded file from disk search: {found_path}{RESET}")
        save_config(found_path)
        return found_path

    # Fallback
    default_path = os.path.join(os.getcwd(), EXCEL_FILE)
    print(f"{RED}Excel file not found. Creating new one at: {default_path}{RESET}")
    save_config(default_path)
    return default_path

EXCEL_PATH = get_excel_path()

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)

def save_to_excel(row_data):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append(row_data)
    wb.save(EXCEL_PATH)

def delete_from_excel(values):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    found = False
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def normalize(val):
            return str(val).strip().replace('\n', '').replace('\r', '')

        compare_values = [normalize(val) for val in values]
        row_values = [normalize(cell.value) for cell in row]

        compare_len = min(len(row_values), len(compare_values))
        if row_values[:compare_len] == compare_values[:compare_len]:
            ws.delete_rows(idx)
            found = True
            print("Successfully Deleted Row")
            break

    wb.save(EXCEL_PATH)
    return found

def get_all_applications():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def update_excel_row(old_values, new_values):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value if cell.value is not None else "" for cell in row]
        if row_values == [v if v is not None else "" for v in old_values]:
            for idx, val in enumerate(new_values):
                row[idx].value = val
            break
    wb.save(EXCEL_PATH)
