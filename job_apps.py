import tkinter as tk
from tkinter import messagebox, ttk
import tkinter.font as tkFont
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import signal
import sys
import webbrowser

EXCEL_FILE = "job_applications.xlsx"
HEADERS = ["Date Applied", "Job Title", "Company", "Location", "Job/Req #", "Link"]
root = None
last_deleted_row = None
undo_btn = None
confirm_btn = None
last_edited_row = None
last_edited_item_id = None
undo_edit_btn = None
confirm_edit_btn = None

PRIMARY_BG = "#2c3e50"
SECONDARY_BG = "#34495e"
TEXT_COLOR = "#ecf0f1"
BUTTON_BG = "#3498db"
BUTTON_FG = "#ffffff"

# handle Ctrl + C exiting
def signal_handler(sig, frame):
    print("\nCtrl+C pressed, exiting...")
    if root is not None:
        root.quit()
    else:
        sys.exit(0)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def parse_job_info(url):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(options=options)

    try:
        driver.get(url)
        time.sleep(5)

        try:
            job_title = "Unknown"
            title_selectors = [
                "//h1",
                "//h2",
                "//*[contains(@class, 'jobTitle')]",
                "//*[contains(@class, 'job-title')]",
                "//*[contains(@class, 'title')]",
                "//*[contains(@id, 'jobTitle')]",
                "//*[contains(@id, 'job-title')]",
                "//div[contains(text(), 'Job Title')]/following-sibling::*[1]",
            ]

            for selector in title_selectors:
                try:
                    elem = driver.find_element(By.XPATH, selector)
                    text = elem.text.strip()
                    if text:
                        job_title = text
                        break
                except:
                    continue
        except:
            job_title = "Unknown"

        try:
            company = driver.find_element(By.XPATH, "//meta[@property='og:site_name']").get_attribute("content").strip()
        except:
            company = "Unknown"

        location = "Unknown"
        elements = driver.find_elements(By.XPATH,
            "//*[contains(@class, 'location') or contains(@id, 'location') or contains(text(), 'United States') or contains(text(), 'Remote')]"
        )

        if not elements:
            elements = driver.find_elements(By.XPATH, "//*[contains(text(), ',')]")

        for el in elements:
            text = el.text.strip()
            if not text:
                continue
            # Only accept short, location-like strings
            if (
                2 <= len(text.split()) <= 6 and
                ',' in text and
                not any(x in text.lower() for x in ['apply', 'requirements', 'responsibilities'])
            ):
                location = text
                break

        job_req = "Unknown"
        try:
            elems = driver.find_elements(
                By.XPATH,
                "//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'job id') or "
                "contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'job number') or "
                "contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'requisition id')]"
            )

            for elem in elems:
                parent = elem.find_element(By.XPATH, "..")
                full_text = parent.text.strip()
                new_text = full_text.split()
                job_num = new_text[2]
                if len(full_text) > len(elem.text):
                    job_req = job_num
                    break
        except:
            pass

        return {
            "Job Title": job_title,
            "Company": company,
            "Location": location,
            "Job/Req #": job_req
        }

    except Exception as e:
        print("Error parsing job info:", e)
        return None
    finally:
        driver.quit()

def treeview_sort_column(tree, col, reverse):
    """Sort treeview contents when a column header is clicked."""
    data = [
        (tree.set(item, col), item) for item in tree.get_children('')
    ]

    # Try to convert to numbers if possible for proper numeric sorting
    try:
        data.sort(key=lambda t: float(t[0]), reverse=reverse)
    except ValueError:
        data.sort(reverse=reverse)

    for index, (val, item) in enumerate(data):
        tree.move(item, '', index)

    # Update all headings to remove sort indicators
    for c in tree['columns']:
        heading_text = tree.heading(c)['text']
        if ' ↑' in heading_text or ' ↓' in heading_text:
            heading_text = heading_text.replace(' ↑', '').replace(' ↓', '')
            tree.heading(c, text=heading_text)

    # Add sort indicator to current column
    sort_symbol = ' ↓' if reverse else ' ↑'
    heading_text = tree.heading(col)['text'].replace(' ↑', '').replace(' ↓', '')
    tree.heading(col, text=heading_text + sort_symbol)

    # Reverse sort next time
    tree.heading(col, command=lambda: treeview_sort_column(tree, col, not reverse))

def add_job(url, tree):
    info = parse_job_info(url)
    if not info:
        return

    today = datetime.today().strftime('%Y-%m-%d')

    row_data = [
        today,
        info["Job Title"],
        info["Company"],
        info["Location"],
        info["Job/Req #"],
        url,
    ]

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(row_data)
    wb.save(EXCEL_FILE)

    tree.insert('', tk.END, values=row_data)

def autosize_columns(tree):
    font = tkFont.Font()
    col_widths = []
    for i, col in enumerate(HEADERS):
        max_width = font.measure(col)
        for row_id in tree.get_children():
            val = tree.item(row_id)['values'][i]
            max_width = max(max_width, font.measure(str(val)))
        # Add some padding to width
        col_widths.append(max_width + 25)
        tree.column(col, width=max_width + 25)
    return col_widths

def remove_selected(tree):
    global last_deleted_row
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("No Selection", "Please select a row to remove.")
        return

    values = tree.item(selected[0], 'values')
    last_deleted_row = values
    tree.delete(selected[0])

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    found = False
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def normalize(val):
            return str(val).strip().replace('\n', '').replace('\r', '')

        compare_values = [normalize(val) for val in values]
        row_values = [normalize(cell.value) for cell in row]

        compare_len = min(len(row_values), len(compare_values))
        if row_values[:compare_len] == compare_values[:compare_len]:
            print("Match found -> Deleting row.")
            ws.delete_rows(idx)
            found = True
            break

    if not found:
        print("Full row not found in Excel for deletion.")

    wb.save(EXCEL_FILE)

    # Enable undo/confirm buttons
    undo_btn.config(state='normal')
    confirm_btn.config(state='normal')

def undo_delete(tree):
    global last_deleted_row
    if not last_deleted_row:
        messagebox.showinfo("Undo Delete", "No deleted job to undo.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(last_deleted_row)
    wb.save(EXCEL_FILE)

    tree.insert('', tk.END, values=last_deleted_row)

    # Reset and disable buttons
    last_deleted_row = None
    undo_btn.config(state='disabled')
    confirm_btn.config(state='disabled')

def confirm_deletion():
    global last_deleted_row
    last_deleted_row = None
    undo_btn.config(state='disabled')
    confirm_btn.config(state='disabled')

def undo_edit(tree):
    global last_edited_row, last_edited_item_id
    if not last_edited_row or not last_edited_item_id:
        messagebox.showinfo("Undo Edit", "No edits to undo.")
        return

    # Restore old values in treeview
    tree.item(last_edited_item_id, values=last_edited_row)

    # Restore old values in Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value if cell.value is not None else "" for cell in row]
        if row_values == [v if v is not None else "" for v in tree.item(last_edited_item_id, 'values')]:
            for idx, val in enumerate(last_edited_row):
                row[idx].value = val
            break
    wb.save(EXCEL_FILE)

    # Reset and disable buttons
    last_edited_row = None
    last_edited_item_id = None
    undo_edit_btn.config(state='disabled')
    confirm_edit_btn.config(state='disabled')

def confirm_edit():
    global last_edited_row, last_edited_item_id
    last_edited_row = None
    last_edited_item_id = None
    undo_edit_btn.config(state='disabled')
    confirm_edit_btn.config(state='disabled')

def show_row_details(tree):
    selected = tree.selection()
    if not selected:
        return

    values = tree.item(selected[0], 'values')

    details_win = tk.Toplevel()
    details_win.title("Job Application Details")
    details_win.geometry("600x400")
    details_win.configure(bg=PRIMARY_BG)
    details_win.resizable(False, False)

    # Force to front to allow Escape
    details_win.lift()
    details_win.focus_force()
    details_win.bind("<Escape>", lambda event: details_win.destroy())

    for idx, col in enumerate(HEADERS):
        tk.Label(
            details_win,
            text=f"{col}:",
            font=('Arial', 10, 'bold'),
            bg=PRIMARY_BG,
            fg=TEXT_COLOR
        ).grid(row=idx, column=0, sticky=tk.W, padx=10, pady=5)

        if col == "Link":
            url = values[idx]

            def open_url(event, link=url):
                webbrowser.open_new(link)

            link_label = tk.Label(
                details_win,
                text=url,
                fg="skyblue",
                cursor="hand2",
                bg=PRIMARY_BG,
                wraplength=450,
                justify=tk.LEFT
            )
            link_label.grid(row=idx, column=1, sticky=tk.W, padx=10, pady=5)
            link_label.bind("<Button-1>", open_url)

        else:
            text = tk.Text(
                details_win,
                height=2,
                wrap=tk.WORD,
                width=60,
                font=('Arial', 10),
                bg=SECONDARY_BG,
                fg=TEXT_COLOR,
                relief='flat',
                bd=0
            )
            text.insert(tk.END, values[idx])
            text.config(state='disabled')
            text.grid(row=idx, column=1, sticky=tk.W, padx=10, pady=5)

def edit_selected(tree):
    global last_edited_row, last_edited_item_id, undo_edit_btn, confirm_edit_btn

    selected = tree.selection()
    if not selected:
        messagebox.showwarning("No Selection", "Please select a row to edit.")
        return

    item_id = selected[0]
    values = list(tree.item(item_id, 'values'))

    edit_win = tk.Toplevel()
    edit_win.title("Edit Job Entry")
    edit_win.geometry("600x400")
    edit_win.configure(bg=PRIMARY_BG)
    edit_win.resizable(False, False)

    entries = {}

    def save_changes():
        nonlocal values, item_id

        new_values = []
        for idx, col in enumerate(HEADERS):
            new_val = entries[col].get().strip()
            new_values.append(new_val if new_val else "Unknown")

        # Save old values for undo
        last_edited_row = values.copy()
        last_edited_item_id = item_id

        # Update treeview immediately
        tree.item(item_id, values=new_values)

        # Update Excel
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            row_values = [cell.value if cell.value is not None else "" for cell in row]
            if row_values == [v if v is not None else "" for v in values]:
                for idx, val in enumerate(new_values):
                    row[idx].value = val
                break
        wb.save(EXCEL_FILE)

        # Enable undo/confirm edit buttons
        undo_edit_btn.config(state='normal')
        confirm_edit_btn.config(state='normal')

        # Update globals
        globals()['last_edited_row'] = last_edited_row
        globals()['last_edited_item_id'] = last_edited_item_id

        edit_win.destroy()

    for idx, col in enumerate(HEADERS):
        tk.Label(
            edit_win,
            text=col,
            font=('Arial', 10, 'bold'),
            bg=PRIMARY_BG,
            fg=TEXT_COLOR
        ).grid(row=idx, column=0, padx=10, pady=5, sticky=tk.W)

        entry = tk.Entry(
            edit_win,
            width=50,
            bg=SECONDARY_BG,
            fg=TEXT_COLOR,
            insertbackground=TEXT_COLOR,  # Cursor color
            relief='flat',
            highlightthickness=1,
            highlightbackground=BUTTON_BG,
            highlightcolor=BUTTON_BG,
            font=('Arial', 10)
        )
        entry.grid(row=idx, column=1, padx=10, pady=5)
        entry.insert(0, values[idx])
        entries[col] = entry

    save_btn = tk.Button(
        edit_win,
        text="Save",
        bg=BUTTON_BG,
        fg=BUTTON_FG,
        font=('Arial', 10, 'bold'),
        command=save_changes
    )
    save_btn.grid(row=len(HEADERS), column=0, columnspan=2, pady=15)

    # Focus on first entry
    entries[HEADERS[0]].focus_set()

def main():
    init_excel()

    global root, undo_btn, confirm_btn, undo_edit_btn, confirm_edit_btn

    root = tk.Tk()
    root.title("Job Tracker")
    root.configure(bg=PRIMARY_BG)

    # Set a reasonable minimum window size and make it resizable
    root.minsize(1000, 600)
    root.geometry("1200x700")

    signal.signal(signal.SIGINT, signal_handler)

    style = ttk.Style()
    style.theme_use("default")
    style.configure("Treeview",
                    background="#ecf0f1",
                    foreground="black",
                    rowheight=25,
                    fieldbackground="#ecf0f1")
    style.map("Treeview",
              background=[('selected', '#3498db')],
              foreground=[('selected', 'white')])

    # URL Entry Frame
    frame_top = tk.Frame(root, bg=SECONDARY_BG)
    frame_top.pack(pady=10, fill=tk.X, padx=10)

    tk.Label(frame_top, text="Job Posting URL:", bg=SECONDARY_BG, fg=TEXT_COLOR, font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    url_entry = tk.Entry(frame_top, width=60, font=('Arial', 10))
    url_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

    def on_add():
        url = url_entry.get().strip()
        if not url:
            messagebox.showwarning("Input Error", "Please enter a job URL.")
            return
        add_job(url, tree)
        url_entry.delete(0, tk.END)
        refresh_treeview()

    url_entry.bind("<Return>", lambda event: on_add())
    add_job_btn = tk.Button(frame_top, text="Add Job", bg=BUTTON_BG, fg=BUTTON_FG,
                           font=('Arial', 10, 'bold'), command=on_add)
    add_job_btn.pack(side=tk.RIGHT, padx=5)

    # Search Frame
    frame_search = tk.Frame(root, bg=SECONDARY_BG)
    frame_search.pack(pady=5, fill=tk.X, padx=10)

    search_var = tk.StringVar()
    tk.Label(frame_search, text="Search:", bg=SECONDARY_BG, fg=TEXT_COLOR, font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    search_entry = tk.Entry(frame_search, textvariable=search_var, width=40, font=('Arial', 10))
    search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

    def clear_search():
        search_var.set("")
        refresh_treeview()

    def refresh_treeview(filter_text=None):
        for item in tree.get_children():
            tree.delete(item)
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_vals = list(row)
            if filter_text:
                if not any(filter_text.lower() in str(cell).lower() for cell in row_vals):
                    continue
            tree.insert('', tk.END, values=row_vals)
        root.update_idletasks()

    def do_search(event=None):
        query = search_var.get().strip()
        if query == "":
            refresh_treeview()
        else:
            refresh_treeview(query)

    search_entry.bind("<Return>", do_search)

    # Search buttons frame
    search_btn_frame = tk.Frame(frame_search, bg=SECONDARY_BG)
    search_btn_frame.pack(side=tk.RIGHT, padx=5)

    tk.Button(search_btn_frame, text="Search", bg=BUTTON_BG, fg=BUTTON_FG,
             font=('Arial', 9), command=do_search).pack(side=tk.LEFT, padx=2)
    tk.Button(search_btn_frame, text="Clear", bg=BUTTON_BG, fg=BUTTON_FG,
             font=('Arial', 9), command=clear_search).pack(side=tk.LEFT, padx=2)

    # Create a frame for the treeview and scrollbars
    tree_frame = tk.Frame(root, bg=PRIMARY_BG)
    tree_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=(10, 5))

    # Treeview with scrollbars
    tree = ttk.Treeview(tree_frame, columns=HEADERS, show='headings')
    treeview_sort_column(tree, "Date Applied", True)  # Sort by date, newest first

    # Vertical scrollbar
    v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=v_scrollbar.set)

    # Horizontal scrollbar
    h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
    tree.configure(xscrollcommand=h_scrollbar.set)

    # Grid layout for treeview and scrollbars
    tree.grid(row=0, column=0, sticky='nsew')
    v_scrollbar.grid(row=0, column=1, sticky='ns')
    h_scrollbar.grid(row=1, column=0, sticky='ew')

    # Configure grid weights so treeview expands
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    for col in HEADERS:
        tree.heading(col, text=col, command=lambda c=col: treeview_sort_column(tree, c, False))
        tree.column(col, anchor=tk.W, width=200)

    tree.bind("<Double-1>", lambda event: show_row_details(tree))
    tree.bind("<Delete>", lambda event: remove_selected(tree))

    # Main Control Buttons Frame - organized in a grid layout
    main_btn_frame = tk.Frame(root, bg=PRIMARY_BG)
    main_btn_frame.pack(pady=(5, 10), padx=10, fill=tk.X)

    # Create a grid layout for better organization
    # Row 1: Primary action buttons
    primary_frame = tk.Frame(main_btn_frame, bg=PRIMARY_BG)
    primary_frame.pack(fill=tk.X, pady=2)

    tk.Button(primary_frame, text="Remove Selected Job", bg="#e74c3c", fg="white",
             font=('Arial', 10, 'bold'), width=18,
             command=lambda: remove_selected(tree)).pack(side=tk.LEFT, padx=5)

    tk.Button(primary_frame, text="Edit Selected Job", bg=BUTTON_BG, fg=BUTTON_FG,
             font=('Arial', 10, 'bold'), width=18,
             command=lambda: edit_selected(tree)).pack(side=tk.LEFT, padx=5)

    # Row 2: Delete action buttons (undo/confirm)
    delete_frame = tk.Frame(main_btn_frame, bg=PRIMARY_BG)
    delete_frame.pack(fill=tk.X, pady=1)

    tk.Label(delete_frame, text="Delete Actions:", bg=PRIMARY_BG, fg=TEXT_COLOR,
            font=('Arial', 9, 'italic')).pack(side=tk.LEFT, padx=5)

    undo_btn = tk.Button(delete_frame, text="Undo Delete", bg="#27ae60", fg="white",
                        font=('Arial', 9, 'bold'), width=15, state='disabled',
                        command=lambda: undo_delete(tree))
    undo_btn.pack(side=tk.LEFT, padx=5)

    confirm_btn = tk.Button(delete_frame, text="Confirm Delete", bg="#e74c3c", fg="white",
                           font=('Arial', 9, 'bold'), width=15, state='disabled',
                           command=confirm_deletion)
    confirm_btn.pack(side=tk.LEFT, padx=5)

    # Row 3: Edit action buttons (undo/confirm)
    edit_frame = tk.Frame(main_btn_frame, bg=PRIMARY_BG)
    edit_frame.pack(fill=tk.X, pady=1)

    tk.Label(edit_frame, text="Edit Actions:", bg=PRIMARY_BG, fg=TEXT_COLOR,
            font=('Arial', 9, 'italic')).pack(side=tk.LEFT, padx=5)

    undo_edit_btn = tk.Button(edit_frame, text="Undo Edit", bg="#27ae60", fg="white",
                             font=('Arial', 9, 'bold'), width=15, state='disabled',
                             command=lambda: undo_edit(tree))
    undo_edit_btn.pack(side=tk.LEFT, padx=5)

    confirm_edit_btn = tk.Button(edit_frame, text="Confirm Edit", bg="#e74c3c", fg="white",
                                font=('Arial', 9, 'bold'), width=15, state='disabled',
                                command=confirm_edit)
    confirm_edit_btn.pack(side=tk.LEFT, padx=5)

    # Keyboard shortcuts info - more compact
    info_frame = tk.Frame(root, bg=PRIMARY_BG)
    info_frame.pack(pady=(0, 5), padx=10, fill=tk.X)

    shortcuts_text = "• Double-click row for details • Delete key to remove • Enter to add/search"
    tk.Label(info_frame, text=shortcuts_text, bg=PRIMARY_BG, fg="#95a5a6",
            font=('Arial', 8), justify=tk.CENTER).pack()

    refresh_treeview()

    def poll():
        root.after(100, poll)
    poll()

    def on_close():
        global last_deleted_row, last_edited_row, last_edited_item_id
        last_deleted_row = None
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()
    print("Application closed.")

if __name__ == "__main__":
    main()
